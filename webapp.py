"""Telegram Mini App (CRM) — operator chat paneli uchun web server (aiohttp).

Bot bilan bitta jarayonda ishlaydi, bitta bazaga ulanadi. Operator mini app'da:
login/parol -> chat ro'yxati -> yozishma -> mijozga yuborish (bot orqali) -> chatni o'chirish.
"""
import os
import html as _htm
import json
import time
import hmac
import base64
import hashlib
import asyncio
import logging
from urllib.parse import parse_qsl

from aiohttp import web
from aiogram.types import BufferedInputFile

from config import BOT_TOKEN, WEBAPP_URL, AVATAR_DIR, MEDIA_CACHE, ADMIN_IDS
from database import queries as q
import locales as loc

logger = logging.getLogger("bot")
_HTML = os.path.join(os.path.dirname(__file__), "webapp", "operator.html")
_ADMIN_HTML = os.path.join(os.path.dirname(__file__), "webapp", "admin.html")


# ---------------- Auth: Telegram WebApp initData ----------------
def _check_init(init_data: str, token: str):
    """initData imzosini token bilan tekshiradi. To'g'ri bo'lsa parsed dict qaytaradi."""
    try:
        parsed = dict(parse_qsl(init_data, strict_parsing=True))
    except Exception:
        return None
    got = parsed.pop("hash", None)
    if not got:
        return None
    data_check = "\n".join(f"{k}={parsed[k]}" for k in sorted(parsed))
    secret = hmac.new(b"WebAppData", token.encode(), hashlib.sha256).digest()
    calc = hmac.new(secret, data_check.encode(), hashlib.sha256).hexdigest()
    return parsed if hmac.compare_digest(calc, got) else None


async def _all_tokens():
    """Asosiy bot + barcha operator bot tokenlari (mini app istalgan botdan ochilishi mumkin)."""
    tokens = [BOT_TOKEN]
    try:
        for b in await q.list_operator_bots(only_enabled=True):
            if b["token"]:
                tokens.append(b["token"])
    except Exception:
        pass
    return tokens


async def _auth_user(request, body=None):
    """initData'dan (header/body/query) haqiqiy Telegram foydalanuvchisini qaytaradi.
    Imzo to'g'ri bo'lmasa None — bu faqat onlayn bog'lash uchun, majburiy emas."""
    init_data = (request.headers.get("X-Init-Data", "")
                 or (body or {}).get("_init", "")
                 or request.query.get("_init", ""))
    if not init_data:
        return None
    for tok in await _all_tokens():
        parsed = _check_init(init_data, tok)
        if parsed:
            raw = parsed.get("user")
            if raw:
                try:
                    return json.loads(raw)
                except Exception:
                    return None
    return None


def _sign(operator_id) -> str:
    """Operator sessiya tokeni (login/parol bilan olinadi, server siri bilan imzolanadi)."""
    return hmac.new(BOT_TOKEN.encode(), f"op-session:{operator_id}".encode(),
                    hashlib.sha256).hexdigest()


async def _auth_op(request, data):
    """Operatorni imzolangan token orqali tasdiqlaydi (login/parol bilan olingan)."""
    try:
        operator_id = int(data.get("operator_id"))
    except (TypeError, ValueError):
        return None, None
    token = str(data.get("token", ""))
    if not token or not hmac.compare_digest(_sign(operator_id), token):
        return None, None
    op = await q.get_operator(operator_id)
    if not op or op["status"] != "active":
        return None, None
    # Ish vaqti tekshiruvi: vaqt tugagach mini app sessiyasi ham yopiladi
    try:
        from utils import operator_in_hours
        within, _ws, _we = operator_in_hours(op)
        if not within:
            return None, None
    except Exception:
        pass
    return op, None


def _json(data, status=200):
    return web.json_response(data, status=status)


# ---------------- Sahifa ----------------
async def index(request):
    if os.path.exists(_HTML):
        # Kesh o'chirilgan: har ochilganda eng yangi dizayn yuklanadi
        return web.FileResponse(_HTML, headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache", "Expires": "0"})
    return web.Response(text="Mini app fayli topilmadi.", status=404)


async def health(request):
    return web.Response(text="ok")


# ---------------- API: login ----------------
async def api_login(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    login = str(body.get("login", "")).strip()
    password = str(body.get("password", "")).strip()
    op = await q.get_operator_by_login(login)
    if not op or op["password_hash"] != q.hash_password(password):
        return _json({"ok": False, "error": "Login yoki parol xato"}, 200)
    if op["status"] != "active":
        return _json({"ok": False, "error": "Hisob bloklangan"}, 200)
    # Ish vaqtidan tashqarida mini appga kirib bo'lmaydi
    from utils import operator_in_hours
    within, ws, we = operator_in_hours(op)
    if not within:
        return _json({"ok": False, "error": f"Hozir ish vaqtingiz emas.\n"
                                            f"Ish vaqtingiz: {ws}–{we}. "
                                            f"Faqat shu oraliqda kira olasiz."}, 200)
    # Telegram foydalanuvchisini (imzo to'g'ri bo'lsa) operatorga bog'laymiz — online bo'ladi
    user = await _auth_user(request, body)
    if user:
        try:
            await q.login_operator(op["id"], user["id"])
        except Exception:
            pass
    return _json({"ok": True, "operator_id": op["id"], "name": op["name"],
                  "token": _sign(op["id"])})


# ---------------- API: chatlar ----------------
def _ct_label(ct):
    return {"photo": "📷 rasm", "video": "🎥 video", "document": "📄 hujjat", "voice": "🎤 ovoz",
            "sticker": "🎭 stiker", "animation": "🎞 GIF", "location": "📍 lokatsiya"}.get(ct, "📎 media")


async def api_chats(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False, "error": "auth"}, 401)
    rows = await q.op_chats(op["id"])
    chats = []
    for r in rows:
        preview = r["last_text"] or (_ct_label(r["last_ct"]) if r["last_ct"] else "")
        chats.append({
            "order_id": r["id"],
            "name": r["full_name"] or "—",
            "phone": r["phone"] or "",
            "status": r["status"],
            "incoming": r["status"] == "new",
            "mine": r["operator_id"] == op["id"],
            "reply": r["last_sender"] == "client",   # javob kutyapti
            "preview": (preview or "")[:60],
            "time": (r["last_at"] or r["created_at"] or "")[11:16],
        })
    return _json({"ok": True, "chats": chats})


# ---------------- API: yozishma ----------------
async def api_messages(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False, "error": "auth"}, 401)
    try:
        order_id = int(request.query.get("order_id"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "order_id"}, 400)
    order = await q.get_order(order_id)
    if not order:
        return _json({"ok": False, "error": "not found"}, 404)
    user = await q.get_user(order["user_id"])
    msgs = await q.order_messages(order_id)
    out = []
    for m in msgs:
        out.append({
            "mid": m["id"],
            "sender": m["sender"],
            "own": m["sender"] == "operator",
            "type": m["content_type"] or "text",
            "text": m["text"] or "",
            "file_id": m["file_id"] or "",
            "tgid": m["tg_msg_id"] or 0,   # reply (iqtibos) uchun — mijoz xabari IDsi
            "cmid": m["client_msg_id"] or 0,   # mijoz chatidagi ID — o'chirish/tahrirlash mumkin
            "time": (m["created_at"] or "")[11:16],
        })
    uname = user["username"] if user and "username" in user.keys() else ""
    return _json({"ok": True, "order_id": order_id, "status": order["status"],
                  "client": {"name": user["full_name"] if user else "—",
                             "phone": user["phone"] if user else "",
                             "username": uname or ""},
                  "messages": out})


# ---------------- API: media (rasm/ovoz) ko'rsatish ----------------
async def api_file(request):
    op, _ = await _auth_op(request, request.query)
    if not op and not await _auth_admin(request, request.query):
        return web.Response(status=401, text="auth")
    fid = request.query.get("fid", "")
    kind = request.query.get("kind", "")
    if not fid:
        return web.Response(status=400)
    ctype = {"voice": "audio/ogg", "audio": "audio/ogg", "video": "video/mp4",
             "sticker": "image/webp", "document": "application/octet-stream"}.get(kind, "image/jpeg")
    cache_path = os.path.join(MEDIA_CACHE, hashlib.sha256(fid.encode()).hexdigest())
    # 1) Keshdan (Telegram'ga qayta so'rov yubormaymiz — egress tejaladi)
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "rb") as fh:
                return web.Response(body=fh.read(), content_type=ctype,
                                    headers={"Cache-Control": "public, max-age=604800"})
        except Exception:
            pass
    # 2) Telegram'dan bir marta yuklab, keshga saqlaymiz.
    # file_id asosiy botniki ham, operator botiniki ham bo'lishi mumkin — hammasini sinaymiz
    from utils import cbot
    import botreg
    candidates = [b for b in ([cbot()] + list(botreg.all_operator_bots().values())) if b]
    if not candidates:
        return web.Response(status=503)
    raw = None
    for b in candidates:
        try:
            f = await b.get_file(fid)
            buf = await b.download_file(f.file_path)
            raw = buf.read()
            break
        except Exception:
            continue
    if raw is None:
        return web.Response(status=404, text="not found")
    try:
        os.makedirs(MEDIA_CACHE, exist_ok=True)
        with open(cache_path, "wb") as fh:
            fh.write(raw)
    except Exception:
        pass
    return web.Response(body=raw, content_type=ctype,
                        headers={"Cache-Control": "public, max-age=604800"})


# ---------------- API: yuborish ----------------
async def api_send(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, user = await _auth_op(request, body)
    if not op:
        return _json({"ok": False, "error": "auth"}, 401)
    try:
        order_id = int(body.get("order_id"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "order_id"}, 400)
    text = str(body.get("text", "")).strip()
    media_kind = body.get("media_kind")          # 'photo' | 'voice' | None
    media_data = body.get("media_data")          # base64 (dataURL bo'lishi mumkin)
    if not text and not (media_kind and media_data):
        return _json({"ok": False, "error": "empty"}, 400)
    order = await q.get_order(order_id)
    if not order or order["status"] not in ("new", "in_progress"):
        return _json({"ok": False, "error": "yopilgan"}, 200)

    # yangi bo'lsa — avval qabul qilamiz (o'zimizga biriktiramiz)
    if order["status"] == "new" and not order["operator_id"]:
        if await q.claim_order(order_id, op["id"]):
            await q.set_operator_availability(op["id"], "busy")
    await q.set_operator_active_order(op["id"], order_id)

    from utils import cbot, post_operator_to_channel
    client = cbot()
    if not client:
        return _json({"ok": False, "error": "bot tayyor emas"}, 200)
    uid = order["user_id"]
    clang = await q.get_lang(uid)
    # Reply (iqtibos): operator mijozning aniq xabariga javob bersa — Telegramda o'sha xabarga tirkaladi
    rkw = {}
    try:
        rtg = int(body.get("reply_tgid") or 0)
        if rtg:
            rkw = {"reply_to_message_id": rtg, "allow_sending_without_reply": True}
    except (TypeError, ValueError):
        pass
    try:
        if media_kind and media_data:
            raw = base64.b64decode(str(media_data).split(",")[-1])
            cap = f"👨‍⚕️ {op['name']}" + (f": {_htm.escape(text)}" if text else "")
            if media_kind == "document":
                fname = str(body.get("media_name") or "hujjat")[:64] or "hujjat"
                sent = await client.send_document(uid, BufferedInputFile(raw, fname),
                                                  caption=cap, **rkw)
                fid = sent.document.file_id
                await q.add_message(order_id, "operator", "document",
                                    text or fname, fid, None, client_msg_id=sent.message_id)
                await post_operator_to_channel(client, order, op["name"], content_type="document",
                                               file_id=fid, src_bot=client, text=text or fname)
            elif media_kind == "photo":
                sent = await client.send_photo(uid, BufferedInputFile(raw, "photo.jpg"),
                                               caption=cap, **rkw)
                fid = sent.photo[-1].file_id
                await q.add_message(order_id, "operator", "photo", text or None, fid, None,
                                    client_msg_id=sent.message_id)
                await post_operator_to_channel(client, order, op["name"], content_type="photo",
                                               file_id=fid, src_bot=client, text=text or "")
            else:  # voice — voice -> audio -> document zanjiri (Telegram formatga qarab)
                mime = str(body.get("media_mime", "")).lower()
                ext = "ogg" if "ogg" in mime else ("webm" if "webm" in mime else
                                                   ("mp4" if "mp4" in mime else "audio"))
                fid, ctype = None, "voice"
                for kind in ("voice", "audio", "document"):
                    try:
                        if kind == "voice":
                            snt = await client.send_voice(uid, BufferedInputFile(raw, "voice.ogg"), **rkw)
                            fid, ctype = snt.voice.file_id, "voice"
                        elif kind == "audio":
                            snt = await client.send_audio(uid, BufferedInputFile(raw, "audio." + ext), **rkw)
                            fid, ctype = snt.audio.file_id, "audio"
                        else:
                            snt = await client.send_document(uid, BufferedInputFile(raw, "voice." + ext),
                                                             caption="🎤 ovozli xabar", **rkw)
                            fid, ctype = snt.document.file_id, "document"
                        break
                    except Exception:
                        continue
                if not fid:
                    return _json({"ok": False, "error": "ovoz yuborilmadi (format qo'llanmadi)"}, 200)
                await q.add_message(order_id, "operator", ctype, None, fid, None,
                                    client_msg_id=snt.message_id)
                await post_operator_to_channel(client, order, op["name"], content_type=ctype,
                                               file_id=fid, src_bot=client, text="🎤 ovozli xabar")
        else:
            snt = await client.send_message(uid, loc.t("operator_reply", clang, name=op["name"],
                                                       text=_htm.escape(text)), **rkw)
            await q.add_message(order_id, "operator", "text", text, None, None,
                                client_msg_id=snt.message_id)
            await post_operator_to_channel(client, order, op["name"], text=text)
    except Exception:
        return _json({"ok": False, "error": "mijozga yuborilmadi (bloklagan bo'lishi mumkin)"}, 200)
    return _json({"ok": True})


# ---------------- API: qabul qilish ----------------
async def api_accept(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, _ = await _auth_op(request, body)
    if not op:
        return _json({"ok": False, "error": "auth"}, 401)
    try:
        order_id = int(body.get("order_id"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "order_id"}, 400)
    from utils import cbot
    from handlers.operator import do_accept
    ok, err = await do_accept(cbot(), op, order_id, op["telegram_id"] or 0)
    return _json({"ok": ok, "error": err})


# ---------------- API: yakunlash ----------------
async def api_close(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, _ = await _auth_op(request, body)
    if not op:
        return _json({"ok": False, "error": "auth"}, 401)
    try:
        order_id = int(body.get("order_id"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "order_id"}, 400)
    from utils import cbot
    from handlers.operator import _finish_with_rating
    await _finish_with_rating(cbot(), order_id, f"operator:{op['id']}")
    return _json({"ok": True})


# ---------------- API: chatni o'chirish (yashirish) ----------------
async def api_hide(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, _ = await _auth_op(request, body)
    if not op:
        return _json({"ok": False, "error": "auth"}, 401)
    try:
        order_id = int(body.get("order_id"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "order_id"}, 400)
    await q.hide_chat(op["id"], order_id)
    return _json({"ok": True})


# ---------------- API: profil ----------------
async def api_profile(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False, "error": "auth"}, 401)
    s = await q.operator_stats(op["id"])
    incoming = await q.new_count()
    return _json({"ok": True, "name": op["name"], "login": op["login"],
                  "availability": op["availability"], "incoming": incoming,
                  "has_avatar": os.path.exists(os.path.join(AVATAR_DIR, f"{op['id']}.jpg")),
                  "stats": {"accepted": s["accepted"], "done": s["done"],
                            "today_done": s["today_done"], "rating": s["avg_rating"],
                            "rated": s["rated_count"]}})


async def api_status(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, _ = await _auth_op(request, body)
    if not op:
        return _json({"ok": False}, 401)
    new = "busy" if op["availability"] == "free" else "free"
    await q.set_operator_availability(op["id"], new)
    return _json({"ok": True, "availability": new})


async def api_avatar(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return web.Response(status=401)
    path = os.path.join(AVATAR_DIR, f"{op['id']}.jpg")
    if os.path.exists(path):
        return web.FileResponse(path, headers={"Cache-Control": "no-cache"})
    return web.Response(status=404)


async def api_avatar_upload(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, _ = await _auth_op(request, body)
    if not op:
        return _json({"ok": False}, 401)
    data = body.get("data")
    if not data:
        return _json({"ok": False, "error": "empty"}, 400)
    try:
        raw = base64.b64decode(str(data).split(",")[-1])
    except Exception:
        return _json({"ok": False, "error": "bad"}, 400)
    os.makedirs(AVATAR_DIR, exist_ok=True)
    with open(os.path.join(AVATAR_DIR, f"{op['id']}.jpg"), "wb") as f:
        f.write(raw)
    return _json({"ok": True})


# ---------------- API: mijozlar ----------------
async def api_clients(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False}, 401)
    search = request.query.get("q") or None
    rows = await q.my_clients(op["id"], search)
    out = [{"tg": r["telegram_id"], "name": r["full_name"] or "—", "phone": r["phone"] or "",
            "cnt": r["cnt"], "last_order": r["last_order"]} for r in rows]
    return _json({"ok": True, "total": len(out), "clients": out})


async def api_client_open(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False}, 401)
    try:
        tg = int(request.query.get("tg"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "tg"}, 400)
    oid = await q.last_order_of(tg)
    if not oid:
        # Murojaat bo'lmasa — yangi suhbat (murojaat) yaratamiz, shu operatorга biriktiramiz
        user = await q.get_user(tg)
        if not user:
            return _json({"ok": False, "error": "Mijoz topilmadi"})
        oid = await q.create_order(tg, user["branch_id"], "text")
        await q.claim_order(oid, op["id"])
        await q.set_operator_active_order(op["id"], oid)
        await q.set_operator_availability(op["id"], "busy")
    return _json({"ok": True, "order_id": oid})


async def api_branches(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False}, 401)
    bs = await q.list_branches()
    return _json({"ok": True, "branches": [{"id": b["id"], "name": b["name"]} for b in bs]})


async def api_newcount(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False}, 401)
    return _json({"ok": True, "count": await q.new_count()})


# ---------------- API: CRM kanali (murojaatlar tushadi) ----------------
async def api_channel(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False}, 401)
    rows = await q.channel_feed()
    items = []
    for r in rows:
        items.append({
            "order_id": r["id"], "name": r["full_name"] or "—", "phone": r["phone"] or "",
            "uname": r["username"] or "", "branch": r["branch"] or "",
            "time": (r["created_at"] or "")[11:16],
            "text": r["first_text"] or "", "file_id": r["first_file"] or "",
            "ftype": r["first_ct"] or ""})
    return _json({"ok": True, "count": len(items), "items": items})


async def api_channel_accept(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, _ = await _auth_op(request, body)
    if not op:
        return _json({"ok": False}, 401)
    try:
        order_id = int(body.get("order_id"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "order_id"}, 400)
    order = await q.get_order(order_id)
    if not order:
        return _json({"ok": False, "error": "topilmadi"}, 404)
    if order["status"] != "new" or not await q.claim_order(order_id, op["id"]):
        return _json({"ok": False, "error": "Boshqa operator qabul qildi"})
    await q.set_operator_active_order(op["id"], order_id)
    await q.set_operator_availability(op["id"], "busy")
    # operator botlaridagi bildirishnomalarni o'chiramiz (boshqalardan yo'qoladi)
    import botreg
    for n in await q.order_notifs(order_id):
        nb = botreg.get_operator_bot(n["bot_id"])
        if nb:
            try:
                await nb.delete_message(n["chat_id"], n["message_id"])
            except Exception:
                pass
    await q.clear_order_notifs(order_id)
    # haqiqiy Telegram kanalidagi kartani yangilaymiz (Jarayonda + kim qabul qildi)
    from utils import cbot, update_group_card
    from config import OPERATORS_GROUP_ID
    client = cbot()
    if client and order["group_msg_id"] and OPERATORS_GROUP_ID:
        try:
            await client.unpin_chat_message(OPERATORS_GROUP_ID, message_id=order["group_msg_id"])
        except Exception:
            pass
        try:
            await update_group_card(client, order_id)
        except Exception:
            pass
    if client:
        clang = await q.get_lang(order["user_id"])
        try:
            await client.send_message(order["user_id"], loc.t("accept_notify", clang))
        except Exception:
            pass
    return _json({"ok": True, "order_id": order_id})


# ---------------- API: chat buyruqlari (/10daqiqa, /filialtanlatish ...) ----------------
async def api_cmd(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, _ = await _auth_op(request, body)
    if not op:
        return _json({"ok": False}, 401)
    try:
        order_id = int(body.get("order_id"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "order_id"}, 400)
    cmd = body.get("cmd")
    arg = body.get("arg")
    order = await q.get_order(order_id)
    if not order:
        return _json({"ok": False, "error": "topilmadi"}, 404)
    from utils import cbot, post_operator_to_channel
    import keyboards as kb
    client = cbot()
    if not client:
        return _json({"ok": False, "error": "bot tayyor emas"}, 200)
    uid = order["user_id"]
    clang = await q.get_lang(uid)
    try:
        if cmd == "autoclose":
            import asyncio
            from handlers.operator import _auto_close_task, AUTO_CLOSE_MIN
            asyncio.create_task(_auto_close_task(client, order_id, q.now(), op["telegram_id"] or 0))
            await client.send_message(uid, loc.t("auto_close_warning", clang, min=AUTO_CLOSE_MIN))
            return _json({"ok": True, "info": f"{AUTO_CLOSE_MIN} daqiqada avto-yakunlash yoqildi"})

        if cmd == "askbranch":
            branches = await q.list_branches()
            await client.send_message(uid, loc.t("op_ask_branch", clang),
                                      reply_markup=kb.op_ask_branch_kb(branches, order_id, clang))
            return _json({"ok": True, "info": "Mijozga filial tanlash so'rovi yuborildi"})

        if cmd == "sendbranch":
            b = await q.get_branch(int(arg))
            if not b:
                return _json({"ok": False, "error": "filial topilmadi"})
            text = (f"🏥 <b>{b['name']}</b>\n📍 {b['address'] or '—'}\n"
                    f"📞 {b['phone'] or '—'}\n🕐 Ish vaqti: {b['open_time']}–{b['close_time']}")
            has = b["lat"] is not None and b["lon"] is not None
            markup = kb.branch_directions_kb(b["lat"], b["lon"]) if has else None
            if b["photo_file_id"]:
                await client.send_photo(uid, b["photo_file_id"], caption=text, reply_markup=markup)
            else:
                await client.send_message(uid, text, reply_markup=markup)
            if has:
                await client.send_venue(uid, latitude=b["lat"], longitude=b["lon"],
                                        title=b["name"], address=b["address"] or "")
            return _json({"ok": True, "info": "Filial ma'lumoti yuborildi"})

        if cmd == "bill":
            billtext = str(arg or "").strip()
            media_kind = body.get("media_kind")
            media_data = body.get("media_data")
            sticker_id = body.get("sticker_id")
            cap = loc.t("bill_to_client", clang, id=order_id, bill=_htm.escape(billtext))
            if media_kind == "photo" and media_data:
                raw = base64.b64decode(str(media_data).split(",")[-1])
                sent = await client.send_photo(uid, BufferedInputFile(raw, "bill.jpg"), caption=cap)
                fid = sent.photo[-1].file_id
                await q.set_order_bill(order_id, billtext, fid)
                await q.add_message(order_id, "operator", "photo",
                                    f"🧾 Hisob-kitob: {billtext}", fid, None)
                await post_operator_to_channel(client, order, op["name"], content_type="photo",
                                               file_id=fid, src_bot=client,
                                               text=f"🧾 Hisob-kitob: {billtext}")
            elif media_kind == "voice" and media_data:
                raw = base64.b64decode(str(media_data).split(",")[-1])
                try:
                    sent = await client.send_voice(uid, BufferedInputFile(raw, "voice.ogg"))
                    fid, ct2 = sent.voice.file_id, "voice"
                except Exception:
                    sent = await client.send_audio(uid, BufferedInputFile(raw, "bill_audio.ogg"))
                    fid, ct2 = sent.audio.file_id, "audio"
                await client.send_message(uid, cap)
                await q.set_order_bill(order_id, billtext or "🎤 ovozli hisob-kitob", None)
                await q.add_message(order_id, "operator", ct2, None, fid, None)
                await post_operator_to_channel(client, order, op["name"], content_type=ct2,
                                               file_id=fid, src_bot=client,
                                               text="🧾 Hisob-kitob (ovozli)")
            elif sticker_id:
                tpl = await q.get_template(int(sticker_id))
                if not tpl or not tpl["sticker"]:
                    return _json({"ok": False, "error": "stiker topilmadi"})
                await client.send_sticker(uid, tpl["sticker"])
                await client.send_message(uid, cap)
                await q.set_order_bill(order_id, billtext or "🎭 stiker", None)
                await q.add_message(order_id, "operator", "sticker", None, tpl["sticker"], None)
                await post_operator_to_channel(client, order, op["name"], content_type="sticker",
                                               file_id=tpl["sticker"], src_bot=client,
                                               text="🧾 Hisob-kitob")
            else:
                await q.set_order_bill(order_id, billtext, None)
                await client.send_message(uid, cap)
                await q.add_message(order_id, "operator", "text",
                                    f"🧾 Hisob-kitob: {billtext}", None, None)
                await post_operator_to_channel(client, order, op["name"],
                                               text=f"🧾 Hisob-kitob: {billtext}")
            return _json({"ok": True, "info": "Hisob-kitob yuborildi"})
    except Exception:
        return _json({"ok": False, "error": "bajarilmadi"}, 200)
    return _json({"ok": False, "error": "noma'lum buyruq"}, 400)


# ---------------- API: sozlamalar ro'yxatlari ----------------
async def api_unfinished(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False}, 401)
    rows = await q.unfinished_orders()
    items = [{"order_id": r["id"], "name": r["full_name"] or "—", "status": r["status"],
              "operator": r["operator"] or "", "mine": r["operator_id"] == op["id"],
              "time": (r["created_at"] or "")[11:16]} for r in rows]
    return _json({"ok": True, "items": items})


async def api_done(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False}, 401)
    rows = await q.orders_by_status("done", op["id"])
    items = []
    for r in rows[:50]:
        u = await q.get_user(r["user_id"])
        items.append({"order_id": r["id"], "name": u["full_name"] if u else "—",
                      "time": (r["closed_at"] or r["created_at"] or "")[:16], "rating": r["rating"] or 0})
    return _json({"ok": True, "items": items})


async def api_rating(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False}, 401)
    rows = await q.operators_rating()
    items = [{"name": r["name"], "score": r["score"], "done": r["done"],
              "rating": r["avg_rating"]} for r in rows]
    return _json({"ok": True, "items": items, "me": op["name"]})


# ---------------- API: stikerlar (admin shablonlari) ----------------
async def api_stickers(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False}, 401)
    tpls = await q.list_templates()
    items = [{"id": t["id"], "file_id": t["sticker"]} for t in tpls if t["sticker"]]
    return _json({"ok": True, "items": items})


async def api_send_sticker(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, _ = await _auth_op(request, body)
    if not op:
        return _json({"ok": False}, 401)
    try:
        order_id = int(body.get("order_id")); tid = int(body.get("sticker_id"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "arg"}, 400)
    order = await q.get_order(order_id)
    tpl = await q.get_template(tid)
    if not order or not tpl or not tpl["sticker"]:
        return _json({"ok": False, "error": "topilmadi"}, 404)
    if order["status"] not in ("new", "in_progress"):
        return _json({"ok": False, "error": "yopilgan"})
    if order["status"] == "new" and not order["operator_id"]:
        if await q.claim_order(order_id, op["id"]):
            await q.set_operator_availability(op["id"], "busy")
    await q.set_operator_active_order(op["id"], order_id)
    from utils import cbot, post_operator_to_channel
    client = cbot()
    if not client:
        return _json({"ok": False, "error": "bot tayyor emas"})
    try:
        snt = await client.send_sticker(order["user_id"], tpl["sticker"])
        await q.add_message(order_id, "operator", "sticker", None, tpl["sticker"], None,
                            client_msg_id=snt.message_id)
        await post_operator_to_channel(client, order, op["name"],
                                       content_type="sticker", file_id=tpl["sticker"], src_bot=client)
    except Exception:
        return _json({"ok": False, "error": "yuborilmadi"})
    return _json({"ok": True})


# ---------------- API: xabarni o'chirish / tahrirlash ----------------
async def api_msg_del(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, _ = await _auth_op(request, body)
    if not op:
        return _json({"ok": False}, 401)
    try:
        mid = int(body.get("mid"))
    except (TypeError, ValueError):
        return _json({"ok": False}, 400)
    row = await q.get_message(mid)
    if not row or row["sender"] != "operator" or not row["client_msg_id"]:
        return _json({"ok": False, "error": "Bu xabarni o'chirib bo'lmaydi"})
    order = await q.get_order(row["order_id"])
    if not order or order["operator_id"] != op["id"]:
        return _json({"ok": False, "error": "Bu sizning suhbatingiz emas"})
    from utils import cbot
    client = cbot()
    if client:
        try:
            await client.delete_message(order["user_id"], row["client_msg_id"])
        except Exception:
            return _json({"ok": False, "error": "Mijozda o'chirib bo'lmadi (48 soatdan oshgan)"})
    await q.delete_message_row(mid)
    return _json({"ok": True})


async def api_msg_edit(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, _ = await _auth_op(request, body)
    if not op:
        return _json({"ok": False}, 401)
    try:
        mid = int(body.get("mid"))
    except (TypeError, ValueError):
        return _json({"ok": False}, 400)
    new_text = str(body.get("text", "")).strip()
    if not new_text:
        return _json({"ok": False, "error": "Matn bo'sh"})
    row = await q.get_message(mid)
    if (not row or row["sender"] != "operator" or not row["client_msg_id"]
            or (row["content_type"] or "text") != "text"):
        return _json({"ok": False, "error": "Faqat o'z matnli xabaringizni tahrirlash mumkin"})
    order = await q.get_order(row["order_id"])
    if not order or order["operator_id"] != op["id"]:
        return _json({"ok": False, "error": "Bu sizning suhbatingiz emas"})
    from utils import cbot
    client = cbot()
    if not client:
        return _json({"ok": False, "error": "bot tayyor emas"})
    clang = await q.get_lang(order["user_id"])
    try:
        await client.edit_message_text(
            loc.t("operator_reply", clang, name=op["name"], text=_htm.escape(new_text)),
            chat_id=order["user_id"], message_id=row["client_msg_id"])
    except Exception:
        return _json({"ok": False, "error": "Tahrirlab bo'lmadi (48 soatdan oshgan)"})
    await q.update_message_text(mid, new_text)
    return _json({"ok": True})


# ---------------- API: eslatma ----------------
async def api_remind(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, _ = await _auth_op(request, body)
    if not op:
        return _json({"ok": False}, 401)
    try:
        order_id = int(body.get("order_id"))
        minutes = int(body.get("minutes"))
    except (TypeError, ValueError):
        return _json({"ok": False}, 400)
    minutes = max(1, min(minutes, 7 * 24 * 60))
    from datetime import timedelta
    from config import now_local
    remind_at = (now_local() + timedelta(minutes=minutes)).strftime("%Y-%m-%d %H:%M:%S")
    await q.add_reminder(op["id"], order_id, remind_at, str(body.get("note", "")).strip()[:200])
    label = f"{minutes} daqiqadan" if minutes < 60 else f"{minutes // 60} soatdan"
    return _json({"ok": True, "info": f"Eslatma qo'yildi — {label} keyin botda xabar keladi"})


# ---------------- API: shaxsiy statistika (7 kun) ----------------
async def api_mystats(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False}, 401)
    from datetime import timedelta
    from config import now_local
    daily = await q.my_daily_done(op["id"], 7)
    wd = ["Du", "Se", "Ch", "Pa", "Ju", "Sh", "Ya"]
    days = []
    for i in range(6, -1, -1):
        d = now_local() - timedelta(days=i)
        key = d.strftime("%Y-%m-%d")
        days.append({"label": wd[d.weekday()], "c": daily.get(key, 0)})
    return _json({"ok": True, "days": days})


# ---------------- API: matnli tayyor javoblar ----------------
async def api_templates(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False}, 401)
    tpls = await q.list_templates()
    items = [{"id": t["id"], "text": t["text"]} for t in tpls if t["text"]]
    return _json({"ok": True, "items": items})


# ---------------- API: mijoz izohi ----------------
async def api_note(request):
    op, _ = await _auth_op(request, request.query)
    if not op:
        return _json({"ok": False}, 401)
    try:
        order_id = int(request.query.get("order_id"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "order_id"}, 400)
    order = await q.get_order(order_id)
    if not order:
        return _json({"ok": False, "error": "topilmadi"}, 404)
    return _json({"ok": True, "note": await q.get_client_note(order["user_id"])})


async def api_note_save(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    op, _ = await _auth_op(request, body)
    if not op:
        return _json({"ok": False}, 401)
    try:
        order_id = int(body.get("order_id"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "order_id"}, 400)
    order = await q.get_order(order_id)
    if not order:
        return _json({"ok": False, "error": "topilmadi"}, 404)
    await q.set_client_note(order["user_id"], str(body.get("note", "")).strip())
    return _json({"ok": True})


# ================================================================
#                        ADMIN MINI APP
# ================================================================
def _admin_sign(tg_id) -> str:
    return hmac.new(BOT_TOKEN.encode(), f"admin-session:{tg_id}".encode(),
                    hashlib.sha256).hexdigest()


async def _auth_admin(request, data):
    """Admin sessiya tokenini tekshiradi (avto-login orqali olingan)."""
    try:
        tg = int(data.get("admin_id"))
    except (TypeError, ValueError):
        return None
    token = str(data.get("token", ""))
    if tg not in ADMIN_IDS:
        return None
    if not token or not hmac.compare_digest(_admin_sign(tg), token):
        return None
    return tg


async def admin_index(request):
    if os.path.exists(_ADMIN_HTML):
        return web.FileResponse(_ADMIN_HTML, headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache", "Expires": "0"})
    return web.Response(text="Admin mini app topilmadi.", status=404)


async def api_admin_login(request):
    """Avto-login: Telegram initData imzosi tekshiriladi, ADMIN_IDS'da bo'lsa kiradi."""
    try:
        body = await request.json()
    except Exception:
        body = {}
    user = await _auth_user(request, body)
    if not user:
        return _json({"ok": False, "error": "Telegram tekshiruvi o'tmadi. "
                                            "Asosiy botdagi Mini app tugmasidan oching."})
    if user["id"] not in ADMIN_IDS:
        return _json({"ok": False, "error": "Siz admin emassiz."})
    return _json({"ok": True, "admin_id": user["id"],
                  "name": user.get("first_name") or "Admin",
                  "token": _admin_sign(user["id"])})


def _period_start_str(period: str) -> str:
    from datetime import timedelta
    from config import now_local
    n = now_local()
    if period == "today":
        return n.strftime("%Y-%m-%d 00:00:00")
    if period == "week":
        return (n - timedelta(days=n.weekday())).strftime("%Y-%m-%d 00:00:00")
    if period == "month":
        return n.strftime("%Y-%m-01 00:00:00")
    if period == "year":
        return n.strftime("%Y-01-01 00:00:00")
    return "0000-01-01 00:00:00"


async def api_admin_dash(request):
    tg = await _auth_admin(request, request.query)
    if not tg:
        return _json({"ok": False, "error": "auth"}, 401)
    period = request.query.get("period", "week")
    since = _period_start_str(period)
    rep = await q.period_report(since)
    hours = await q.hourly_load(since)
    live = await q.live_stats()
    series = await q.series_counts(since, "month" if period == "year" else "day")
    rating, rated = await q.period_rating(since)
    ops = []
    for o in live["per_op"]:
        st = "offline" if not o["telegram_id"] else ("busy" if o["availability"] == "busy" else "free")
        ops.append({"id": o["id"], "name": o["name"], "state": st,
                    "cnt": o["cnt"], "done_today": o["done_today"]})
    # Hozir javob kutayotgan murojaatlar (necha daqiqadan beri)
    from datetime import datetime as _dt
    from config import now_local as _nl
    waiting = []
    for w in await q.waiting_orders():
        try:
            age = int((_nl() - _dt.strptime(w["created_at"], "%Y-%m-%d %H:%M:%S")).total_seconds() // 60)
        except Exception:
            age = 0
        waiting.append({"id": w["id"], "name": w["full_name"] or "—", "mins": max(age, 0)})
    # Davr ichida eng ko'p murojaat yuborgan mijozlar
    tc_rows, _tc_total = await q.top_clients(since, 6, 0)
    topclients = [{"tg": r["telegram_id"], "name": r["full_name"] or "—",
                   "phone": r["phone"] or "", "cnt": r["cnt"]} for r in tc_rows]
    # Operatorlar taqqoslash (davr bo'yicha yakunlar) + filial kesimi
    oprep = await q.operators_report(since)
    opstats = sorted([{"name": o["name"], "done": o["done"]} for o in oprep if o["done"]],
                     key=lambda x: x["done"], reverse=True)[:8]
    branches = [{"name": b["name"], "cnt": b["cnt"]} for b in await q.branch_counts(since)]
    return _json({"ok": True, "period": period, "waiting": waiting, "topclients": topclients,
                  "opstats": opstats, "branches": branches,
                  "kpi": {"total": rep["total"], "new": rep["new"], "prog": rep["prog"],
                          "done": rep["done"], "canceled": rep["canceled"],
                          "resp": rep["resp"], "resol": rep["resol"],
                          "rating": rating, "rated": rated, "online": live["online"]},
                  "series": [{"d": s["d"], "total": s["total"], "done": s["done"] or 0} for s in series],
                  "hours": [{"h": h, "c": hours.get(h, 0)} for h in range(24)],
                  "ops": ops})


# ---------------- Admin: murojaatlar ro'yxati + yozishma ----------------
async def api_admin_orders(request):
    if not await _auth_admin(request, request.query):
        return _json({"ok": False, "error": "auth"}, 401)
    try:
        page = int(request.query.get("page", "0") or 0)
    except ValueError:
        page = 0
    search = request.query.get("q") or None
    status = request.query.get("status") or None
    period = request.query.get("period") or None
    since = _period_start_str(period) if period and period != "all" else None
    rows, total = await q.orders_page(20, page * 20, search, status, since)
    items = [{"id": r["id"], "name": r["full_name"] or "—", "phone": r["phone"] or "",
              "status": r["status"], "operator": r["operator"] or "",
              "time": (r["created_at"] or "")[5:16], "rating": r["rating"] or 0} for r in rows]
    return _json({"ok": True, "total": total, "page": page, "items": items})


async def api_admin_msgs(request):
    if not await _auth_admin(request, request.query):
        return _json({"ok": False, "error": "auth"}, 401)
    try:
        order_id = int(request.query.get("order_id"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "order_id"}, 400)
    order = await q.get_order(order_id)
    if not order:
        return _json({"ok": False, "error": "topilmadi"}, 404)
    user = await q.get_user(order["user_id"])
    op = await q.get_operator(order["operator_id"]) if order["operator_id"] else None
    msgs = await q.order_messages(order_id)
    out = [{"own": m["sender"] == "operator", "type": m["content_type"] or "text",
            "text": m["text"] or "", "file_id": m["file_id"] or "",
            "time": (m["created_at"] or "")[11:16]} for m in msgs]
    return _json({"ok": True, "order_id": order_id, "status": order["status"],
                  "operator": op["name"] if op else "",
                  "rating": order["rating"] or 0,
                  "client": {"name": user["full_name"] if user else "—",
                             "phone": user["phone"] if user else ""},
                  "messages": out})


async def api_admin_close(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    tg = await _auth_admin(request, body)
    if not tg:
        return _json({"ok": False, "error": "auth"}, 401)
    try:
        order_id = int(body.get("order_id"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "order_id"}, 400)
    order = await q.get_order(order_id)
    if not order or order["status"] not in ("new", "in_progress"):
        return _json({"ok": False, "error": "Bu murojaat allaqachon yopilgan"})
    from utils import cbot
    from handlers.operator import _finish_with_rating
    await _finish_with_rating(cbot(), order_id, f"admin:{tg}")
    return _json({"ok": True})


# ---------------- Admin: mijozlar ----------------
async def api_admin_clients(request):
    if not await _auth_admin(request, request.query):
        return _json({"ok": False, "error": "auth"}, 401)
    try:
        page = int(request.query.get("page", "0") or 0)
    except ValueError:
        page = 0
    search = request.query.get("q") or None
    rows, total = await q.users_page(20, page * 20, search)
    items = [{"tg": r["telegram_id"], "name": r["full_name"] or "—", "phone": r["phone"] or "",
              "branch": r["branch"] or "", "cnt": r["cnt"],
              "last": (r["last_at"] or "")[:10]} for r in rows]
    return _json({"ok": True, "total": total, "page": page, "items": items})


async def api_admin_client(request):
    if not await _auth_admin(request, request.query):
        return _json({"ok": False, "error": "auth"}, 401)
    try:
        tg_ = int(request.query.get("tg"))
    except (TypeError, ValueError):
        return _json({"ok": False, "error": "tg"}, 400)
    u = await q.user_full(tg_)
    if not u:
        return _json({"ok": False, "error": "topilmadi"}, 404)
    orders = await q.orders_by_user(tg_)
    note = await q.get_client_note(tg_)
    return _json({"ok": True,
                  "client": {"tg": tg_, "name": u["full_name"] or "—", "phone": u["phone"] or "",
                             "branch": u["branch"] or "", "reg": (u["registered_at"] or "")[:10],
                             "blocked": u["status"] == "blocked", "note": note},
                  "orders": [{"id": o["id"], "status": o["status"],
                              "date": (o["created_at"] or "")[:10],
                              "rating": o["rating"] or 0} for o in orders[:30]]})


# ================= Admin: OPERATORLAR BOSHQARUVI =================
async def api_admin_ops(request):
    if not await _auth_admin(request, request.query):
        return _json({"ok": False}, 401)
    ops = await q.list_operators()
    bots = {b["id"]: b["username"] for b in await q.list_operator_bots()}
    items = []
    for o in ops:
        items.append({"id": o["id"], "name": o["name"], "login": o["login"],
                      "active": o["status"] == "active",
                      "online": bool(o["telegram_id"]),
                      "avail": o["availability"],
                      "ws": o["work_start"], "we": o["work_end"],
                      "bot": ("@" + bots[o["bot_id"]]) if o["bot_id"] in bots else "asosiy bot"})
    return _json({"ok": True, "items": items})


async def api_admin_op_save(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    name = str(body.get("name", "")).strip()
    login = str(body.get("login", "")).strip()
    password = str(body.get("password", "")).strip()
    ws = str(body.get("ws", "08:00")).strip() or "08:00"
    we = str(body.get("we", "23:00")).strip() or "23:00"
    oid = body.get("id")
    if not name or not login:
        return _json({"ok": False, "error": "Ism va login bo'sh bo'lmasin"})
    existing = await q.get_operator_by_login(login)
    if existing and (not oid or existing["id"] != int(oid)):
        return _json({"ok": False, "error": "Bu login band"})
    if oid:
        oid = int(oid)
        await q.update_operator(oid, "name", name)
        await q.update_operator(oid, "login", login)
        await q.update_operator(oid, "work_start", ws)
        await q.update_operator(oid, "work_end", we)
        if password:
            await q.update_operator_password(oid, password)
            await q.clear_operator_session(oid)
    else:
        if not password:
            return _json({"ok": False, "error": "Parol kiriting"})
        await q.add_operator(name, login, password, bot_id=None)
    return _json({"ok": True})


async def api_admin_op_toggle(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    oid = int(body.get("id"))
    op = await q.get_operator(oid)
    if not op:
        return _json({"ok": False, "error": "topilmadi"}, 404)
    new = "inactive" if op["status"] == "active" else "active"
    await q.update_operator(oid, "status", new)
    if new == "inactive" and op["telegram_id"]:
        await q.logout_operator(op["telegram_id"], op["bot_id"])
    return _json({"ok": True, "active": new == "active"})


async def api_admin_op_del(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    await q.delete_operator(int(body.get("id")))
    return _json({"ok": True})


async def api_admin_op_detail(request):
    if not await _auth_admin(request, request.query):
        return _json({"ok": False}, 401)
    try:
        oid = int(request.query.get("id"))
    except (TypeError, ValueError):
        return _json({"ok": False}, 400)
    op = await q.get_operator(oid)
    if not op:
        return _json({"ok": False, "error": "topilmadi"}, 404)
    s = await q.operator_stats(oid)
    recent = await q.orders_by_operator(oid)
    return _json({"ok": True,
                  "op": {"name": op["name"], "login": op["login"],
                         "active": op["status"] == "active", "online": bool(op["telegram_id"]),
                         "ws": op["work_start"], "we": op["work_end"]},
                  "stats": s,
                  "recent": [{"id": r["id"], "name": r["full_name"] or "—", "status": r["status"],
                              "date": (r["created_at"] or "")[:10], "rating": r["rating"] or 0}
                             for r in recent[:15]]})


async def api_admin_opbots(request):
    if not await _auth_admin(request, request.query):
        return _json({"ok": False}, 401)
    bots = await q.list_operator_bots()
    items = []
    for b in bots:
        ops = await q.operators_by_bot(b["id"])
        items.append({"id": b["id"], "username": b["username"], "title": b["title"],
                      "enabled": bool(b["enabled"]), "ops": len(ops),
                      "logins": ", ".join(o["login"] for o in ops)})
    return _json({"ok": True, "items": items})


async def api_admin_opbot_add(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    token = str(body.get("token", "")).strip()
    login = str(body.get("login", "")).strip()
    password = str(body.get("password", "")).strip()
    if not token or not login or not password:
        return _json({"ok": False, "error": "Token, login va parolni kiriting"})
    if await q.get_operator_bot_by_token(token):
        return _json({"ok": False, "error": "Bu bot allaqachon qo'shilgan"})
    if await q.get_operator_by_login(login):
        return _json({"ok": False, "error": "Bu login band"})
    from aiogram import Bot as _TgBot
    try:
        tb = _TgBot(token)
        me = await tb.get_me()
        await tb.session.close()
    except Exception:
        return _json({"ok": False, "error": "Token noto'g'ri yoki bot topilmadi"})
    bot_id = await q.add_operator_bot(token, me.username, me.first_name)
    await q.add_operator(me.first_name, login, password, bot_id=bot_id)
    import botreg
    try:
        await botreg.start_operator_bot(bot_id, token)
    except Exception:
        pass
    return _json({"ok": True, "username": me.username})


async def api_admin_opbot_toggle(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    bid = int(body.get("id"))
    b = await q.get_operator_bot(bid)
    if not b:
        return _json({"ok": False}, 404)
    import botreg
    new = not b["enabled"]
    await q.set_operator_bot_enabled(bid, new)
    try:
        if new:
            await botreg.start_operator_bot(bid, b["token"])
        else:
            await botreg.stop_operator_bot(bid)
    except Exception:
        pass
    return _json({"ok": True, "enabled": new})


async def api_admin_opbot_del(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    bid = int(body.get("id"))
    import botreg
    try:
        await botreg.stop_operator_bot(bid)
    except Exception:
        pass
    await q.delete_operator_bot(bid)
    return _json({"ok": True})


async def api_admin_lowratings(request):
    if not await _auth_admin(request, request.query):
        return _json({"ok": False}, 401)
    rows = await q.low_rated_orders()
    return _json({"ok": True, "items": [
        {"id": r["id"], "rating": r["rating"], "feedback": r["feedback"] or "",
         "name": r["full_name"] or "—", "operator": r["operator"] or "—",
         "date": (r["closed_at"] or "")[:10]} for r in rows]})


# ---------------- Admin: mijozga yozish + o'tkazish ----------------
async def api_admin_send(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    tg = await _auth_admin(request, body)
    if not tg:
        return _json({"ok": False}, 401)
    try:
        order_id = int(body.get("order_id"))
    except (TypeError, ValueError):
        return _json({"ok": False}, 400)
    text = str(body.get("text", "")).strip()
    if not text:
        return _json({"ok": False, "error": "Matn bo'sh"})
    order = await q.get_order(order_id)
    if not order or order["status"] not in ("new", "in_progress"):
        return _json({"ok": False, "error": "Murojaat yopilgan"})
    from utils import cbot, post_operator_to_channel
    client = cbot()
    if not client:
        return _json({"ok": False, "error": "bot tayyor emas"})
    try:
        snt = await client.send_message(order["user_id"],
                                        f"👨‍💼 <b>Admin:</b>\n{_htm.escape(text)}")
    except Exception:
        return _json({"ok": False, "error": "Mijozga yuborilmadi"})
    await q.add_message(order_id, "operator", "text", f"👨‍💼 Admin: {text}", None, None,
                        client_msg_id=snt.message_id)
    try:
        await post_operator_to_channel(client, order, "Admin", text=text)
    except Exception:
        pass
    return _json({"ok": True})


async def api_admin_transfer(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    tg = await _auth_admin(request, body)
    if not tg:
        return _json({"ok": False}, 401)
    try:
        order_id = int(body.get("order_id"))
        newop_id = int(body.get("operator_id"))
    except (TypeError, ValueError):
        return _json({"ok": False}, 400)
    order = await q.get_order(order_id)
    new_op = await q.get_operator(newop_id)
    if not order or not new_op:
        return _json({"ok": False, "error": "topilmadi"}, 404)
    if order["status"] not in ("new", "in_progress"):
        return _json({"ok": False, "error": "Murojaat yopilgan"})
    prev_id = order["operator_id"]
    if order["status"] == "new":
        await q.claim_order(order_id, newop_id)
    else:
        await q.assign_order(order_id, newop_id)
    await q.set_operator_active_order(newop_id, order_id)
    await q.set_operator_availability(newop_id, "busy")
    import botreg
    from utils import cbot, update_group_card
    client = cbot()
    if prev_id and prev_id != newop_id:
        prev = await q.get_operator(prev_id)
        await q.set_operator_active_order(prev_id, None)
        await q.set_operator_availability(prev_id, "free")
        if prev and prev["telegram_id"]:
            pb = (botreg.get_operator_bot(prev["bot_id"]) if prev["bot_id"] else client) or client
            try:
                await pb.send_message(prev["telegram_id"],
                                      f"ℹ️ Murojaat #{order_id} admin tomonidan boshqa operatorga o'tkazildi.")
            except Exception:
                pass
    if new_op["telegram_id"]:
        nb = (botreg.get_operator_bot(new_op["bot_id"]) if new_op["bot_id"] else client) or client
        try:
            await nb.send_message(new_op["telegram_id"],
                                  f"🔄 Admin sizga murojaat #{order_id} ni biriktirdi. "
                                  f"CRM yoki botda oching.")
        except Exception:
            pass
    if client:
        try:
            await update_group_card(client, order_id)
        except Exception:
            pass
    return _json({"ok": True, "info": f"#{order_id} → {new_op['name']} ga o'tkazildi"})


# ---------------- Admin: mijozni bloklash / to'liq o'chirish / izoh ----------------
async def api_admin_client_block(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    try:
        tg_ = int(body.get("tg"))
    except (TypeError, ValueError):
        return _json({"ok": False}, 400)
    block = bool(body.get("block"))
    await q.set_user_status(tg_, "blocked" if block else "active")
    return _json({"ok": True, "blocked": block})


async def api_admin_client_del(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    try:
        tg_ = int(body.get("tg"))
    except (TypeError, ValueError):
        return _json({"ok": False}, 400)
    n = await q.delete_client_full(tg_)
    return _json({"ok": True, "orders_removed": n})


async def api_admin_note_save(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    try:
        tg_ = int(body.get("tg"))
    except (TypeError, ValueError):
        return _json({"ok": False}, 400)
    await q.set_client_note(tg_, str(body.get("note", "")).strip())
    return _json({"ok": True})


# ================= Admin: SOZLAMALAR (3-bosqich) =================
async def api_admin_branches(request):
    if not await _auth_admin(request, request.query):
        return _json({"ok": False}, 401)
    bs = await q.list_branches()
    return _json({"ok": True, "items": [
        {"id": b["id"], "name": b["name"], "address": b["address"] or "",
         "phone": b["phone"] or "", "open": b["open_time"], "close": b["close_time"],
         "has_loc": b["lat"] is not None} for b in bs]})


async def api_admin_branch_save(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    name = str(body.get("name", "")).strip()
    if not name:
        return _json({"ok": False, "error": "Nomi bo'sh"})
    addr = str(body.get("address", "")).strip()
    phone = str(body.get("phone", "")).strip()
    open_t = str(body.get("open", "08:00")).strip() or "08:00"
    close_t = str(body.get("close", "23:00")).strip() or "23:00"
    bid = body.get("id")
    if bid:
        for f, v in (("name", name), ("address", addr), ("phone", phone),
                     ("open_time", open_t), ("close_time", close_t)):
            await q.update_branch(int(bid), f, v)
    else:
        await q.add_branch(name, addr, phone, open_time=open_t, close_time=close_t)
    return _json({"ok": True})


async def api_admin_branch_del(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    await q.delete_branch(int(body.get("id")))
    return _json({"ok": True})


async def api_admin_faqs(request):
    if not await _auth_admin(request, request.query):
        return _json({"ok": False}, 401)
    fs = await q.list_faqs()
    return _json({"ok": True, "items": [
        {"id": f["id"], "title": f["title"], "answer": f["answer"]} for f in fs]})


async def api_admin_faq_save(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    title = str(body.get("title", "")).strip()
    answer = str(body.get("answer", "")).strip()
    if not title or not answer:
        return _json({"ok": False, "error": "Sarlavha va javob bo'sh bo'lmasin"})
    fid = body.get("id")
    if fid:
        await q.update_faq(int(fid), title, answer)
    else:
        await q.add_faq(title, answer)
    return _json({"ok": True})


async def api_admin_faq_del(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    await q.delete_faq(int(body.get("id")))
    return _json({"ok": True})


async def api_admin_tpls(request):
    if not await _auth_admin(request, request.query):
        return _json({"ok": False}, 401)
    ts = await q.list_templates()
    return _json({"ok": True, "items": [
        {"id": t["id"], "text": t["text"] or "", "sticker": bool(t["sticker"])} for t in ts]})


async def api_admin_tpl_add(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    text = str(body.get("text", "")).strip()
    if not text:
        return _json({"ok": False, "error": "Matn bo'sh"})
    await q.add_template(text)
    return _json({"ok": True})


async def api_admin_tpl_del(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    await q.delete_template(int(body.get("id")))
    return _json({"ok": True})


async def api_admin_settings(request):
    if not await _auth_admin(request, request.query):
        return _json({"ok": False}, 401)
    return _json({"ok": True,
                  "work_start": await q.get_setting("work_start", "08:00"),
                  "work_end": await q.get_setting("work_end", "23:00"),
                  "op_work_start": await q.get_setting("op_work_start", "08:00"),
                  "op_work_end": await q.get_setting("op_work_end", "23:00"),
                  "escalate_min": await q.get_setting("escalate_min", "5"),
                  "contact_text": await q.get_setting("contact_text", "")})


async def api_admin_settings_save(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    if not await _auth_admin(request, body):
        return _json({"ok": False}, 401)
    for k in ("work_start", "work_end", "op_work_start", "op_work_end",
              "escalate_min", "contact_text"):
        if k in body and str(body[k]).strip() != "":
            await q.set_setting(k, str(body[k]).strip())
    return _json({"ok": True})


async def api_admin_broadcast(request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    tg = await _auth_admin(request, body)
    if not tg:
        return _json({"ok": False}, 401)
    text = str(body.get("text", "")).strip()
    if not text:
        return _json({"ok": False, "error": "Matn bo'sh"})
    target = body.get("target", "all")
    branch_id = body.get("branch_id")
    from utils import cbot
    client = cbot()
    if not client:
        return _json({"ok": False, "error": "bot tayyor emas"})
    if target == "active":
        users = await q.all_users(only_active=True)
    elif target == "branch" and branch_id:
        users = await q.all_users(branch_id=int(branch_id))
    else:
        users = await q.all_users()
    # Rasm bo'lsa: birinchi yuborishda yuklab, keyin file_id bilan (tez)
    media_data = body.get("media_data")
    raw = None
    if media_data:
        try:
            raw = base64.b64decode(str(media_data).split(",")[-1])
        except Exception:
            raw = None
    sent = failed = 0
    fid = None
    for u in users:
        try:
            if raw:
                if fid:
                    await client.send_photo(u["telegram_id"], fid, caption=_htm.escape(text))
                else:
                    s = await client.send_photo(u["telegram_id"],
                                                BufferedInputFile(raw, "elon.jpg"),
                                                caption=_htm.escape(text))
                    fid = s.photo[-1].file_id
            else:
                await client.send_message(u["telegram_id"], _htm.escape(text))
            sent += 1
        except Exception:
            failed += 1
    return _json({"ok": True, "sent": sent, "failed": failed, "total": len(users)})


async def api_admin_excel(request):
    """Excel hisobotni yaratib, adminning Telegram chatiga yuboradi."""
    try:
        body = await request.json()
    except Exception:
        body = {}
    tg = await _auth_admin(request, body)
    if not tg:
        return _json({"ok": False}, 401)
    from utils import cbot, STATUS_LABEL
    client = cbot()
    if not client:
        return _json({"ok": False, "error": "bot tayyor emas"})
    period = body.get("period") or "all"
    since = _period_start_str(period) if period != "all" else None
    plabel = {"today": "Bugun", "week": "Joriy hafta", "month": "Joriy oy",
              "year": "Joriy yil", "all": "Barcha davr"}.get(period, "Barcha davr")
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    rows = await q.all_orders_full(since)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Murojaatlar"
    # Sarlavha
    ws.merge_cells("A1:K1")
    t1 = ws["A1"]
    t1.value = f"Gulnora Farm — Murojaatlar hisoboti · {plabel} · {q.now()[:16]}"
    t1.font = Font(bold=True, size=13, color="FFFFFF")
    t1.fill = PatternFill("solid", fgColor="2F6FB4")
    t1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    # Ustun sarlavhalari
    headers = ["ID", "Mijoz", "Telefon", "Filial", "Operator", "Holat",
               "Tur", "Hisob", "Boshlangan", "Yakunlangan", "Baho"]
    hfill = PatternFill("solid", fgColor="3390EC")
    thin = Side(style="thin", color="C9D4E0")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bd
    ws.row_dimensions[2].height = 20
    st_color = {"new": "B45309", "in_progress": "1D6FBF", "done": "1F7A35", "canceled": "C22F36"}
    zebra = PatternFill("solid", fgColor="F3F7FB")
    for ri, r in enumerate(rows, 3):
        vals = [r["id"], r["full_name"], r["phone"], r["branch"], r["operator"],
                STATUS_LABEL.get(r["status"], r["status"]), r["content_type"],
                r["bill"], r["created_at"], r["closed_at"],
                ("★" * r["rating"]) if r["rating"] else ""]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = bd
            if ri % 2:
                c.fill = zebra
            if ci == 1:
                c.alignment = Alignment(horizontal="center")
            if ci == 6:
                c.font = Font(bold=True, color=st_color.get(r["status"], "000000"), size=10.5)
    for col, w in zip("ABCDEFGHIJK", (7, 22, 17, 20, 16, 15, 11, 18, 17, 17, 10)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    buf = _io.BytesIO()
    wb.save(buf)
    try:
        await client.send_document(
            tg, BufferedInputFile(buf.getvalue(), f"hisobot_{period}.xlsx"),
            caption=f"📥 Murojaatlar hisoboti — {plabel} ({len(rows)} ta yozuv)")
    except Exception:
        return _json({"ok": False, "error": "Botga yuborib bo'lmadi"})
    return _json({"ok": True})


async def api_admin_excel_clients(request):
    """Mijozlar ro'yxatini chiroyli Excel qilib admin botiga yuboradi."""
    try:
        body = await request.json()
    except Exception:
        body = {}
    tg = await _auth_admin(request, body)
    if not tg:
        return _json({"ok": False}, 401)
    from utils import cbot
    client = cbot()
    if not client:
        return _json({"ok": False, "error": "bot tayyor emas"})
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    rows = await q.clients_full()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mijozlar"
    ws.merge_cells("A1:H1")
    t1 = ws["A1"]
    t1.value = f"Gulnora Farm — Mijozlar ro'yxati · {q.now()[:16]} · {len(rows)} ta"
    t1.font = Font(bold=True, size=13, color="FFFFFF")
    t1.fill = PatternFill("solid", fgColor="2F6FB4")
    t1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    headers = ["№", "Ism", "Telefon", "Filial", "Holat", "Ro'yxatdan", "Murojaatlar", "O'rtacha baho"]
    hfill = PatternFill("solid", fgColor="3390EC")
    thin = Side(style="thin", color="C9D4E0")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bd
    zebra = PatternFill("solid", fgColor="F3F7FB")
    for ri, r in enumerate(rows, 3):
        vals = [ri - 2, r["full_name"], r["phone"], r["branch"],
                "Bloklangan" if r["status"] == "blocked" else "Faol",
                (r["registered_at"] or "")[:10], r["cnt"],
                r["avg_r"] if r["avg_r"] else ""]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = bd
            if ri % 2:
                c.fill = zebra
            if ci in (1, 7, 8):
                c.alignment = Alignment(horizontal="center")
            if ci == 5 and r["status"] == "blocked":
                c.font = Font(bold=True, color="C22F36", size=10.5)
    for col, w in zip("ABCDEFGH", (6, 24, 18, 20, 12, 13, 12, 13)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    buf = _io.BytesIO()
    wb.save(buf)
    try:
        await client.send_document(tg, BufferedInputFile(buf.getvalue(), "mijozlar.xlsx"),
                                   caption=f"📥 Mijozlar ro'yxati ({len(rows)} ta)")
    except Exception:
        return _json({"ok": False, "error": "Botga yuborib bo'lmadi"})
    return _json({"ok": True})


# ---------------- Kesh tozalash (disk to'lmasin) ----------------
async def _cache_cleanup_loop():
    while True:
        try:
            if os.path.isdir(MEDIA_CACHE):
                cutoff = time.time() - 7 * 86400
                for name in os.listdir(MEDIA_CACHE):
                    p = os.path.join(MEDIA_CACHE, name)
                    try:
                        if os.path.isfile(p) and os.path.getmtime(p) < cutoff:
                            os.remove(p)
                    except Exception:
                        pass
        except Exception:
            pass
        await asyncio.sleep(24 * 3600)


# ---------------- Server ----------------
def build_app() -> web.Application:
    app = web.Application(client_max_size=25 * 1024 * 1024)
    app.router.add_get("/", index)
    app.router.add_get("/operator", index)
    app.router.add_get("/health", health)
    app.router.add_post("/api/login", api_login)
    app.router.add_get("/api/chats", api_chats)
    app.router.add_get("/api/messages", api_messages)
    app.router.add_get("/api/file", api_file)
    app.router.add_post("/api/send", api_send)
    app.router.add_post("/api/accept", api_accept)
    app.router.add_post("/api/close", api_close)
    app.router.add_post("/api/hide", api_hide)
    app.router.add_get("/api/profile", api_profile)
    app.router.add_post("/api/status", api_status)
    app.router.add_get("/api/avatar", api_avatar)
    app.router.add_post("/api/avatar", api_avatar_upload)
    app.router.add_get("/api/clients", api_clients)
    app.router.add_get("/api/client_open", api_client_open)
    app.router.add_get("/api/branches", api_branches)
    app.router.add_post("/api/cmd", api_cmd)
    app.router.add_get("/api/channel", api_channel)
    app.router.add_post("/api/channel_accept", api_channel_accept)
    app.router.add_get("/api/newcount", api_newcount)
    app.router.add_get("/api/unfinished", api_unfinished)
    app.router.add_get("/api/done", api_done)
    app.router.add_get("/api/rating", api_rating)
    app.router.add_get("/api/stickers", api_stickers)
    app.router.add_post("/api/send_sticker", api_send_sticker)
    app.router.add_get("/api/templates", api_templates)
    app.router.add_get("/api/note", api_note)
    app.router.add_post("/api/note", api_note_save)
    app.router.add_post("/api/msg_del", api_msg_del)
    app.router.add_post("/api/msg_edit", api_msg_edit)
    app.router.add_post("/api/remind", api_remind)
    app.router.add_get("/api/mystats", api_mystats)
    # Admin mini app
    app.router.add_get("/admin", admin_index)
    app.router.add_post("/api/admin/login", api_admin_login)
    app.router.add_get("/api/admin/dash", api_admin_dash)
    app.router.add_get("/api/admin/orders", api_admin_orders)
    app.router.add_get("/api/admin/msgs", api_admin_msgs)
    app.router.add_post("/api/admin/close", api_admin_close)
    app.router.add_get("/api/admin/clients", api_admin_clients)
    app.router.add_get("/api/admin/client", api_admin_client)
    app.router.add_get("/api/admin/branches", api_admin_branches)
    app.router.add_post("/api/admin/branch_save", api_admin_branch_save)
    app.router.add_post("/api/admin/branch_del", api_admin_branch_del)
    app.router.add_get("/api/admin/faqs", api_admin_faqs)
    app.router.add_post("/api/admin/faq_save", api_admin_faq_save)
    app.router.add_post("/api/admin/faq_del", api_admin_faq_del)
    app.router.add_get("/api/admin/tpls", api_admin_tpls)
    app.router.add_post("/api/admin/tpl_add", api_admin_tpl_add)
    app.router.add_post("/api/admin/tpl_del", api_admin_tpl_del)
    app.router.add_get("/api/admin/ops", api_admin_ops)
    app.router.add_post("/api/admin/op_save", api_admin_op_save)
    app.router.add_post("/api/admin/op_toggle", api_admin_op_toggle)
    app.router.add_post("/api/admin/op_del", api_admin_op_del)
    app.router.add_get("/api/admin/op_detail", api_admin_op_detail)
    app.router.add_get("/api/admin/opbots", api_admin_opbots)
    app.router.add_post("/api/admin/opbot_add", api_admin_opbot_add)
    app.router.add_post("/api/admin/opbot_toggle", api_admin_opbot_toggle)
    app.router.add_post("/api/admin/opbot_del", api_admin_opbot_del)
    app.router.add_get("/api/admin/lowratings", api_admin_lowratings)
    app.router.add_post("/api/admin/send", api_admin_send)
    app.router.add_post("/api/admin/transfer", api_admin_transfer)
    app.router.add_post("/api/admin/client_block", api_admin_client_block)
    app.router.add_post("/api/admin/client_del", api_admin_client_del)
    app.router.add_post("/api/admin/note_save", api_admin_note_save)
    app.router.add_post("/api/admin/excel_clients", api_admin_excel_clients)
    app.router.add_get("/api/admin/settings", api_admin_settings)
    app.router.add_post("/api/admin/settings_save", api_admin_settings_save)
    app.router.add_post("/api/admin/broadcast", api_admin_broadcast)
    app.router.add_post("/api/admin/excel", api_admin_excel)
    return app


async def start(port: int):
    runner = web.AppRunner(build_app())
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", port)
    await site.start()
    asyncio.create_task(_cache_cleanup_loop())
    logger.info("🖥 Mini app server: 0.0.0.0:%s  (URL: %s)", port, WEBAPP_URL or "— sozlanmagan")
